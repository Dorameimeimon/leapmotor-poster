#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
晨夕会海报自动生成脚本
使用方法：
1. 将新的Excel文件命名为"晨夕会审核列表.xlsx"并放在同一文件夹
2. 运行此脚本：python auto_generate_poster.py
3. 脚本会自动生成最新的"晨夕会海报.html"
"""

import openpyxl
import json
from datetime import datetime, timedelta
from collections import defaultdict
import os

# 102家店白名单和零售经理映射
WHITELIST_MAP = {
    # 方任昊 - 东莞 (12家)
    "东莞零跑中心塘厦车城店": "方任昊",
    "东莞体验中心东城店": "方任昊",
    "东莞体验中心松山湖店": "方任昊",
    "东莞体验中心石排大道店": "方任昊",
    "东莞零跑中心常朗路店": "方任昊",
    "东莞体验中心高埗车城店": "方任昊",
    "东莞零跑中心虎门车城店": "方任昊",
    "东莞零跑中心寮步车城店": "方任昊",
    "东莞零跑中心莞太路店": "方任昊",
    "东莞体验中心厚街店": "方任昊",
    "东莞零跑中心长安振安东路店": "方任昊",
    "东莞体验中心大岭山镇店": "方任昊",
    # 黄伟峰 - 广州 (10家)
    "广州零跑中心金马汽车城店": "黄伟峰",
    "广州零跑中心亚运大道店": "黄伟峰",
    "广州零跑中心美轮汽车园店": "黄伟峰",
    "广州零跑中心沥滘海心沙店": "黄伟峰",
    "广州体验中心天河汇彩路店": "黄伟峰",
    "广州体验中心天河广汕路店": "黄伟峰",
    "广州体验中心番禺大道店": "黄伟峰",
    "广州体验中心白云机场路店": "黄伟峰",
    "广州零跑中心新塘汽车城店": "黄伟峰",
    "广州零跑中心科学城店": "黄伟峰",
    # 罗捷 - 深圳 (6家)
    "深圳体验中心杰鹏车城店": "罗捷",
    "深圳体验中心龙华中心店": "罗捷",
    "深圳零跑中心光明车城店": "罗捷",
    "深圳零跑中心万国城旗舰店": "罗捷",
    "深圳零跑中心远望车城店": "罗捷",
    "深圳零跑中心宝运达汽车城店": "罗捷",
    # 贾迪赫 - 深圳 (6家)
    "深圳体验中心龙岗信义店": "贾迪赫",
    "深圳体验中心锦龙大道店": "贾迪赫",
    "深圳零跑中心百世汽车城店": "贾迪赫",
    "深圳零跑中心南山嘉进隆店": "贾迪赫",
    "深圳体验中心福田瑞鹏达店": "贾迪赫",
    "深圳零跑中心芙蓉路店": "贾迪赫",
    # 余子恩 - 佛山/江门 (10家)
    "佛山体验中心容桂车城店": "余子恩",
    "佛山零跑中心顺德大良车城店": "余子恩",
    "佛山体验中心北滘车城店": "余子恩",
    "佛山体验中心高明荷香路店": "余子恩",
    "佛山零跑中心三水友好车城店": "余子恩",
    "佛山体验中心狮山汽车城店": "余子恩",
    "江门体验中心开平车城店": "余子恩",
    "江门零跑中心冈州大道店": "余子恩",
    "江门体验中心蓬江华鸿店": "余子恩",
    "江门零跑中心建设路店": "余子恩",
    # 沈祖福 - 佛山 (6家)
    "佛山体验中心海八西路店": "沈祖福",
    "佛山零跑中心南庄汽车城店": "沈祖福",
    "佛山零跑中心禅城国际汽车城店": "沈祖福",
    "佛山零跑中心桂澜路店": "沈祖福",
    "佛山体验中心广佛汽车城店": "沈祖福",
    "佛山零跑中心海八路汽车城店": "沈祖福",
    # 林伟龙 - 中山/珠海 (9家)
    "中山零跑中心小榄菊城大道店": "林伟龙",
    "中山零跑中心金宁汽车城店": "林伟龙",
    "中山体验中心港口店": "林伟龙",
    "中山零跑中心中山六路店": "林伟龙",
    "中山零跑中心彩虹大道店": "林伟龙",
    "珠海体验中心美满车城店": "林伟龙",
    "珠海零跑中心南屏珠海大道店": "林伟龙",
    "珠海体验中心西部车城店": "林伟龙",
    "珠海零跑中心上冲车城店": "林伟龙",
    # 胡浩 - 惠州/汕尾/韶关 (8家)
    "惠州零跑中心金山汽车城店": "胡浩",
    "惠州体验中心惠阳大道店": "胡浩",
    "惠州零跑中心惠南汽车城店": "胡浩",
    "惠州零跑中心仲恺汽车城店": "胡浩",
    "惠州体验中心河南岸车城店": "胡浩",
    "汕尾体验中心汕尾大道店": "胡浩",
    "韶关体验中心浈江大道店": "胡浩",
    "韶关零跑中心沐溪大道店": "胡浩",
    # 熊俊晖 - 广州/清远/汕头/河源 (12家)
    "广州体验中心为正车城店": "熊俊晖",
    "广州零跑中心花都建设路店": "熊俊晖",
    "广州零跑中心白云大道店": "熊俊晖",
    "广州体验中心广花路店": "熊俊晖",
    "广州体验中心花都北站店": "熊俊晖",
    "清远零跑中心港鸿汽车城店": "熊俊晖",
    "清远体验中心奇晟汽车城店": "熊俊晖",
    "汕头体验中心广汕路店": "熊俊晖",
    "汕头零跑中心金凤路店": "熊俊晖",
    "汕头零跑中心汕汾路店": "熊俊晖",
    "河源零跑中心河源大道店": "熊俊晖",
    "河源体验中心坚基购物中心店": "熊俊晖",
    # 庄文迪 - 湛江/肇庆/茂名/阳江 (7家)
    "湛江零跑中心海田车城店": "庄文迪",
    "湛江体验中心粤西车城店": "庄文迪",
    "湛江体验中心瀚龙车城店": "庄文迪",
    "肇庆零跑中心肇庆大道店": "庄文迪",
    "茂名零跑中心茂名大道店": "庄文迪",
    "茂名体验中心茂水路店": "庄文迪",
    "阳江零跑中心溢信汽车城店": "庄文迪",
    # 李博恩 - 三亚/海口 (11家)
    "三亚体验中心迎宾路汽车城店": "李博恩",
    "三亚零跑中心榆亚路店": "李博恩",
    "海口零跑中心琼山大道店": "李博恩",
    "海口体验中心江东店": "李博恩",
    "海口体验中心海口东站店": "李博恩",
    "海口零跑中心海盛路店": "李博恩",
    "海口零跑中心南海大道店（销售）": "李博恩",
    "海口零跑中心南海大道东店": "李博恩",
    "海口体验中心海甸城店": "李博恩",
    "海口体验中心上邦百汇城店": "李博恩",
    "海口零跑中心美安科技城店": "李博恩",
    # 李稳 - 揭阳/梅州/潮州/云浮 (5家)
    "揭阳零跑中心荣通汽车城店": "李稳",
    "揭阳体验中心普宁万泰新天地商场店": "李稳",
    "梅州零跑中心剑英大道店": "李稳",
    "潮州零跑中心潮汕路店": "李稳",
    "云浮体验中心环市中路店": "李稳"
}

MANAGER_ORDER = ["黄伟峰", "熊俊晖", "罗捷", "贾迪赫", "余子恩", "沈祖福", "方任昊", "林伟龙", "胡浩", "李稳", "庄文迪", "李博恩"]

def process_excel(excel_path):
    """从Excel提取会议数据（只统计白名单中的门店）"""
    print(f"正在读取Excel文件: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    # 计算日期范围：昨天往前7天
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)
    start_date = yesterday - timedelta(days=6)  # 昨天往前6天 = 总共7天

    print(f"统计周期: {start_date.strftime('%m/%d')} - {yesterday.strftime('%m/%d')} (昨天往前7天)")

    store_meeting_data = {}
    all_dates = set()
    filtered_count = 0
    filtered_date_count = 0

    for row in range(2, ws.max_row + 1):
        store_name = ws.cell(row=row, column=4).value  # 第4列是门店
        meeting_type = ws.cell(row=row, column=2).value  # 第2列是会议类型
        meeting_date = ws.cell(row=row, column=7).value  # 第7列是会议日期

        if not store_name or not meeting_type or not meeting_date:
            continue

        # 只统计白名单中的门店
        if store_name not in WHITELIST_MAP:
            filtered_count += 1
            continue

        # 解析日期
        if isinstance(meeting_date, str):
            try:
                meeting_date = datetime.strptime(meeting_date, '%Y-%m-%d %H:%M:%S')
            except:
                continue

        if not isinstance(meeting_date, datetime):
            continue

        # 只统计昨天往前7天的数据
        meeting_date_only = meeting_date.date()
        if meeting_date_only < start_date or meeting_date_only > yesterday:
            filtered_date_count += 1
            continue

        day = meeting_date.day
        month = meeting_date.month
        date_key = f"{month}/{day}"
        all_dates.add((month, day))

        # 存储数据
        if store_name not in store_meeting_data:
            store_meeting_data[store_name] = {}

        if day not in store_meeting_data[store_name]:
            store_meeting_data[store_name][day] = {'morning': False, 'evening': False}

        if meeting_type == '晨会':
            store_meeting_data[store_name][day]['morning'] = True
        elif meeting_type == '夕会':
            store_meeting_data[store_name][day]['evening'] = True

    # 确定日期范围
    sorted_dates = sorted(list(all_dates))
    date_labels = [f"{d[0]}/{d[1]}" for d in sorted_dates]
    day_list = [d[1] for d in sorted_dates]

    print(f"日期范围: {date_labels[0]} - {date_labels[-1]}")
    print(f"白名单门店数: {len(WHITELIST_MAP)} 家")
    if filtered_count > 0:
        print(f"已过滤非白名单门店: {filtered_count} 条记录")
    if filtered_date_count > 0:
        print(f"已过滤日期范围外记录: {filtered_date_count} 条记录")

    return store_meeting_data, date_labels, day_list

def build_managers_data(store_meeting_data, date_labels, day_list):
    """构建经理数据"""
    managers_data = []

    for manager in MANAGER_ORDER:
        stores = [store for store, mgr in WHITELIST_MAP.items() if mgr == manager]
        total = len(stores)

        morning_counts = [0] * len(date_labels)
        evening_counts = [0] * len(date_labels)
        store_details = []

        for store in stores:
            morning = []
            evening = []

            for i, day in enumerate(day_list):
                data = store_meeting_data.get(store, {}).get(day, {'morning': False, 'evening': False})

                has_morning = data['morning']
                has_evening = data['evening']

                morning.append(has_morning)
                evening.append(has_evening)

                if has_morning:
                    morning_counts[i] += 1
                if has_evening:
                    evening_counts[i] += 1

            store_details.append({
                'name': store,
                'morning': morning,
                'evening': evening
            })

        # 计算最后一天的完成率
        m_rate = round(morning_counts[-1] / total * 100) if total > 0 else 0
        e_rate = round(evening_counts[-1] / total * 100) if total > 0 else 0

        managers_data.append({
            'name': manager,
            'total': total,
            'morning': morning_counts,
            'evening': evening_counts,
            'morning_rate': f"{m_rate}%",
            'evening_rate': f"{e_rate}%",
            'store_details': store_details
        })

    # 计算总计 - 使用白名单中的总门店数
    total_stores = 102  # 白名单中的总门店数
    total_morning = [sum(m['morning'][i] for m in managers_data) for i in range(len(date_labels))]
    total_evening = [sum(m['evening'][i] for m in managers_data) for i in range(len(date_labels))]

    tm_rate = round(total_morning[-1] / total_stores * 100) if total_stores > 0 else 0
    te_rate = round(total_evening[-1] / total_stores * 100) if total_stores > 0 else 0

    print(f"总计: {total_stores}家门店, 晨会率{tm_rate}%, 夕会率{te_rate}%")

    return managers_data, total_stores, total_morning, total_evening, tm_rate, te_rate

def generate_html(date_labels, managers_data, total_stores, total_morning, total_evening, tm_rate, te_rate):
    """生成HTML文件"""
    print("正在生成HTML文件...")

    # 读取模板或生成完整HTML
    html_template = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=1400">
<title>华南大区晨夕会开展情况</title>
<style>
* { margin: 0; padding: 0; box-sizing: border-box; }
body {
  font-family: -apple-system, "PingFang SC", "Microsoft YaHei", sans-serif;
  background: #f0f2f5;
  display: flex;
  justify-content: center;
  padding: 40px 20px;
}
.poster {
  width: 1400px;
  background: #fff;
  border-radius: 12px;
  overflow: hidden;
  box-shadow: 0 4px 24px rgba(0,0,0,.08);
}
.header {
  background: linear-gradient(135deg, #1a3c6e 0%, #2d5fa0 100%);
  padding: 28px 40px 22px;
  display: flex;
  align-items: center;
  gap: 20px;
}
.header-left { display: flex; align-items: center; gap: 14px; }
.header-dot {
  width: 8px; height: 8px;
  background: #f9a826;
  border-radius: 50%;
  box-shadow: 0 0 8px #f9a82688;
}
.header h1 { color: #fff; font-size: 26px; font-weight: 700; letter-spacing: 2px; }
.header-date {
  color: #bfd4f0; font-size: 14px; margin-left: 10px;
  padding: 4px 14px; background: rgba(255,255,255,.1); border-radius: 20px;
}
.header-stats { margin-left: auto; display: flex; gap: 24px; }
.stat-item { text-align: center; }
.stat-num { color: #f9a826; font-size: 28px; font-weight: 700; }
.stat-label { color: #bfd4f0; font-size: 12px; margin-top: 2px; }
.body { padding: 20px 40px 30px; }

.section-label {
  display: flex; align-items: center; gap: 10px;
  margin: 18px 0 10px;
}
.section-title-wrapper {
  display: flex; justify-content: space-between; align-items: center; flex: 1;
}
.section-title-morning { display: flex; align-items: center; gap: 10px; }
.section-title-evening { display: flex; align-items: center; gap: 10px; }
.section-title-evening .bar { background: #5b8def; }
.section-label .bar { width: 4px; height: 18px; border-radius: 2px; }
.bar.morning { background: #f9a826; }
.bar.evening { background: #5b8def; }
.section-label span { font-size: 16px; font-weight: 600; color: #333; }

.table-wrap { overflow-x: auto; }
table { width: 100%; border-collapse: collapse; font-size: 13px; }
th {
  background: #f5f7fa; padding: 10px 8px; text-align: center;
  font-weight: 600; color: #555; white-space: nowrap;
  border-bottom: 2px solid #e8ecf0; font-size: 13px;
}
td {
  padding: 8px; text-align: center; border-bottom: 1px solid #f0f2f5;
  white-space: nowrap; color: #333; font-size: 13px;
}
.gap-col { width: 20px; background: transparent !important; border: none !important; }
.row-total { background: #e8f5e9; font-weight: 600; }
.rate-tag {
  display: inline-block; padding: 2px 10px; border-radius: 10px;
  font-size: 12px; font-weight: 600;
}
.rate-high { background: #e8f5e9; color: #2e7d32; }
.rate-mid { background: #fff3e0; color: #ef6c00; }
.rate-low { background: #fce4ec; color: #c62828; }
.manager-end-col {
  text-align: center !important; font-weight: 500; color: #333;
  border-left: 2px solid #e8ecf0;
}
.store-section { margin-top: 30px; }
.store-filter {
  display: flex; align-items: center; gap: 10px; margin-bottom: 15px;
}
.store-filter label { font-size: 14px; color: #555; font-weight: 500; }
.store-filter select {
  padding: 6px 12px; font-size: 13px; border: 1px solid #d0d5dd; border-radius: 6px;
  background: #fff; color: #333; cursor: pointer; min-width: 160px;
}
.store-filter select:focus { outline: none; border-color: #2d5fa0; box-shadow: 0 0 0 2px rgba(45,95,160,.15); }
.check-col { padding: 4px 2px !important; }
.check-yes { color: #2e7d32; }
.check-no { color: #bdbdbd; }
</style>
</head>
<body>
<div class="poster">
<div class="header">
  <div class="header-left"><div class="header-dot"></div><h1>华南大区晨夕会开展情况</h1></div>
  <span class="header-date" id="headerDate"></span>
  <div class="header-stats">
    <div class="stat-item"><div class="stat-num" id="morningRate">__TM_RATE__</div><div class="stat-label">昨日晨会率</div></div>
    <div class="stat-item"><div class="stat-num" id="eveningRate">__TE_RATE__</div><div class="stat-label">昨日夕会率</div></div>
  </div>
</div>

<div class="body">

<!-- 经理统计表 -->
<div class="section-label">
  <div class="section-title-wrapper">
    <div class="section-title-morning">
      <div class="bar morning"></div><span>晨会</span>
    </div>
    <div class="section-title-evening">
      <div class="bar"></div><span>夕会</span>
    </div>
  </div>
</div>
<table id="managerTable"></table>

<!-- 门店明细 -->
<div class="store-section">
<div class="section-label">
  <div class="section-title-wrapper">
    <div class="section-title-morning">
      <div class="bar morning"></div><span>晨会</span>
    </div>
    <div class="section-title-evening">
      <div class="bar"></div><span>夕会</span>
    </div>
  </div>
</div>
<div class="store-filter">
  <label>筛选零售高级经理：</label>
  <select id="managerFilter" onchange="filterStores()">
    <option value="all">全部</option>
  </select>
</div>
<div class="table-wrap"><table id="storeTable"></table></div>
</div>

</div>
</div>

<script>
const DATE_LABELS = __DATE_LABELS__;
const TOTAL_STORES = __TOTAL__;
const MANAGERS_DATA = __MANAGERS_DATA__;

function rateClass(rateStr) {
  const v = parseInt(rateStr);
  if (v >= 90) return 'rate-high';
  if (v >= 70) return 'rate-mid';
  return 'rate-low';
}

function renderManagerTable() {
  const table = document.getElementById('managerTable');
  let html = '<thead><tr>';
  html += '<th>零售高级经理</th><th>门店数</th>';
  for (let d of DATE_LABELS) html += '<th>' + d + '</th>';
  html += '<th>晨会率</th><th class="gap-col"></th>';
  for (let d of DATE_LABELS) html += '<th>' + d + '</th>';
  html += '<th>夕会率</th><th>零售高级经理</th>';
  html += '</tr></thead><tbody>';

  let tm = __TOTAL_MORNING__;
  let te = __TOTAL_EVENING__;
  html += '<tr class="row-total">';
  html += '<td>华南</td><td>' + TOTAL_STORES + '</td>';
  for (let v of tm) html += '<td>' + v + '</td>';
  html += '<td><span class="rate-tag ' + rateClass("__TM_RATE__") + '">__TM_RATE__</span></td>';
  html += '<td class="gap-col"></td>';
  for (let v of te) html += '<td>' + v + '</td>';
  html += '<td><span class="rate-tag ' + rateClass("__TE_RATE__") + '">__TE_RATE__</span></td>';
  html += '<td>华南</td>';
  html += '</tr>';

  for (let m of MANAGERS_DATA) {
    html += '<tr>';
    html += '<td>' + m.name + '</td><td>' + m.total + '</td>';
    for (let v of m.morning) html += '<td>' + v + '</td>';
    html += '<td><span class="rate-tag ' + rateClass(m.morning_rate) + '">' + m.morning_rate + '</span></td>';
    html += '<td class="gap-col"></td>';
    for (let v of m.evening) html += '<td>' + v + '</td>';
    html += '<td><span class="rate-tag ' + rateClass(m.evening_rate) + '">' + m.evening_rate + '</span></td>';
    html += '<td class="manager-end-col">' + m.name + '</td>';
    html += '</tr>';
  }
  html += '</tbody>';
  table.innerHTML = html;
}

function renderStoreTable(filterManager) {
  const table = document.getElementById('storeTable');
  let html = '<thead><tr>';
  html += '<th>门店名称</th><th>零售高级经理</th>';
  for (let d of DATE_LABELS) html += '<th>' + d + '</th>';
  html += '<th class="gap-col"></th>';
  for (let d of DATE_LABELS) html += '<th>' + d + '</th>';
  html += '</tr></thead><tbody>';

  for (let m of MANAGERS_DATA) {
    if (filterManager && filterManager !== 'all' && m.name !== filterManager) continue;
    for (let s of m.store_details) {
      html += '<tr>';
      html += '<td>' + s.name + '</td><td>' + m.name + '</td>';
      for (let v of s.morning) html += '<td class="check-col"><span class="' + (v ? 'check-yes' : 'check-no') + '">' + (v ? '✅' : '❌') + '</span></td>';
      html += '<td class="gap-col"></td>';
      for (let v of s.evening) html += '<td class="check-col"><span class="' + (v ? 'check-yes' : 'check-no') + '">' + (v ? '✅' : '❌') + '</span></td>';
      html += '</tr>';
    }
  }
  html += '</tbody>';
  table.innerHTML = html;
}

function filterStores() {
  const filter = document.getElementById('managerFilter').value;
  renderStoreTable(filter);
}

function initManagerFilter() {
  const select = document.getElementById('managerFilter');
  for (let m of MANAGERS_DATA) {
    const option = document.createElement('option');
    option.value = m.name;
    option.textContent = m.name;
    select.appendChild(option);
  }
}

function updateHeaderDate() {
  const yesterday = new Date();
  yesterday.setDate(yesterday.getDate() - 1);
  const month = yesterday.getMonth() + 1;
  const day = yesterday.getDate();
  const dateStr = month + '/' + day;
  document.getElementById('headerDate').textContent = dateStr;
}

updateHeaderDate();
renderManagerTable();
initManagerFilter();
renderStoreTable('all');
</script>
</body>
</html>'''

    # 替换占位符
    html_output = html_template
    html_output = html_output.replace('__DATE_LABELS__', json.dumps(date_labels, ensure_ascii=False))
    html_output = html_output.replace('__TOTAL__', str(total_stores))
    html_output = html_output.replace('__MANAGERS_DATA__', json.dumps(managers_data, ensure_ascii=False))
    html_output = html_output.replace('__TOTAL_MORNING__', json.dumps(total_morning))
    html_output = html_output.replace('__TOTAL_EVENING__', json.dumps(total_evening))
    html_output = html_output.replace('__TM_RATE__', f"{tm_rate}%")
    html_output = html_output.replace('__TE_RATE__', f"{te_rate}%")

    # 写入文件
    with open('晨夕会海报.html', 'w', encoding='utf-8') as f:
        f.write(html_output)

    print("✅ HTML文件已生成: 晨夕会海报.html")
    print(f"   昨日晨会率: {tm_rate}%")
    print(f"   昨日夕会率: {te_rate}%")

def main():
    # 检查命令行参数
    import sys
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
        if not os.path.exists(excel_file):
            print(f"❌ 指定的文件不存在: {excel_file}")
            return
        print(f"正在读取Excel文件: {os.path.basename(excel_file)}")
    else:
        # 查找Excel文件
        excel_file = None
        for file in os.listdir('.'):
            if file.startswith('晨夕会审核列表') and file.endswith('.xlsx'):
                excel_file = file
                break

        if not excel_file:
            print("❌ 未找到Excel文件！请将Excel文件命名为'晨夕会审核列表.xlsx'并放在同一文件夹")
            return

    # 处理Excel
    store_meeting_data, date_labels, day_list = process_excel(excel_file)

    # 构建数据
    managers_data, total_stores, total_morning, total_evening, tm_rate, te_rate = build_managers_data(
        store_meeting_data, date_labels, day_list
    )

    # 生成HTML
    generate_html(date_labels, managers_data, total_stores, total_morning, total_evening, tm_rate, te_rate)

if __name__ == '__main__':
    main()
